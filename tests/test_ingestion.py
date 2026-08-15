from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook

from unisweet_analysis.config import ProjectPaths
from unisweet_analysis.loaders import load_sales


SALES_COLUMNS = [
    "Customer Code",
    "Brand Code",
    "Pack Type",
    "Pack Size",
    "Month",
    "Year",
    "KPI",
    "Values",
]


def test_sales_loader_accepts_a_new_customer_file_without_code_change(tmp_path: Path) -> None:
    (tmp_path / "inputs" / "sales").mkdir(parents=True)
    (tmp_path / "inputs" / "mapping").mkdir()
    (tmp_path / "inputs" / "pnl").mkdir()
    (tmp_path / "inputs" / "market").mkdir()
    (tmp_path / "config").mkdir()
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(SALES_COLUMNS)
    sheet.append([30, "OLI", "PACK", "1.1KG", "Jan", 2025, "Gross Sales Value", 120])
    sheet.append([30, "OLI", "PACK", "1.1KG", "Jan", 2025, "Turnover", 100])
    workbook.save(tmp_path / "inputs" / "sales" / "Cust 30.xlsx")
    settings = {
        "sales_required_columns": SALES_COLUMNS,
        "sales_kpis": {"Gross Sales Value": "GSV", "Turnover": "TURNOVER"},
    }
    (tmp_path / "config" / "settings.json").write_text(json.dumps(settings), encoding="utf-8")
    placeholder = Workbook()
    placeholder.save(tmp_path / "inputs" / "mapping" / "Master Mapping.xlsx")
    placeholder.save(tmp_path / "inputs" / "pnl" / "P&L Table.xlsx")
    placeholder.save(tmp_path / "inputs" / "market" / "Market Report MAT Jan'25.xlsx")

    frame, sources = load_sales(ProjectPaths.from_root(tmp_path), settings)

    assert len(sources) == 1
    assert sources[0].schema_status == "PASS"
    assert len(frame) == 2
    assert frame.iloc[0]["customer_code"] == "00000030"
    assert frame.iloc[0]["reporting_month"].strftime("%Y-%m") == "2025-01"


def test_latest_market_file_is_selected_from_filename(tmp_path: Path) -> None:
    market_dir = tmp_path / "inputs" / "market"
    market_dir.mkdir(parents=True)
    for name in ["Market Report MAT Nov'24.xlsx", "Market Report MAT Jan'25.xlsx"]:
        workbook = Workbook()
        workbook.save(market_dir / name)
    paths = ProjectPaths.from_root(tmp_path)
    assert paths.market_file.name == "Market Report MAT Jan'25.xlsx"
