"""Build one flat Sales master CSV from Customer Sales files and Master Mapping.

This is the only data-processing script in the project. It does not read or
modify the P&L or Market workbooks.
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook


SALES_COLUMNS = [
    "Customer Code",
    "Brand Code",
    "Pack Type",
    "Pack Size",
    "Month",
    "Year",
    "KPI",
    "Values",
]
KPI_MAP = {"Gross Sales Value": "GSV", "Turnover": "TURNOVER"}
MONTH_NUMBERS = {
    month: number
    for number, month in enumerate(
        ["JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"],
        start=1,
    )
}
GRAIN_COLUMNS = [
    "reporting_month",
    "customer_code",
    "brand_code",
    "product_key",
    "pack_type",
    "pack_size",
]


def normalize_code(value: Any, width: int | None = None) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
    else:
        text = str(value).strip()
    return text.zfill(width) if width and text.isdigit() else text


def _sales_sort_key(path: Path) -> tuple[int, int | str]:
    match = re.fullmatch(r"Cust\s+(\d+)", path.stem, re.IGNORECASE)
    return (0, int(match.group(1))) if match else (1, path.name)


def load_customer_sales(project_root: Path) -> tuple[pd.DataFrame, list[Path]]:
    sales_dir = project_root / "inputs" / "sales"
    files = sorted(
        [path for path in sales_dir.glob("Cust *.xlsx") if not path.name.startswith("~$")],
        key=_sales_sort_key,
    )
    if not files:
        raise FileNotFoundError(f"No Customer Sales files found in {sales_dir}")

    frames: list[pd.DataFrame] = []
    schema_errors: list[str] = []
    for path in files:
        frame = pd.read_excel(path, engine="openpyxl", dtype=object)
        valid_name = re.fullmatch(r"Cust\s+(\d+)", path.stem, re.IGNORECASE) is not None
        if frame.columns.tolist() != SALES_COLUMNS or not valid_name:
            schema_errors.append(
                f"{path.name}: expected {SALES_COLUMNS}; received {frame.columns.tolist()}"
            )
            continue
        frame = frame.dropna(how="all").copy()
        frame["source_file"] = path.relative_to(project_root).as_posix()
        frame["source_row"] = frame.index + 2
        frames.append(frame)

    if schema_errors:
        raise ValueError("Invalid Customer Sales schema:\n" + "\n".join(schema_errors))

    raw = pd.concat(frames, ignore_index=True).rename(
        columns={
            "Customer Code": "customer_code",
            "Brand Code": "brand_code",
            "Pack Type": "pack_type",
            "Pack Size": "pack_size",
            "Month": "month_label",
            "Year": "reporting_year",
            "KPI": "metric_label",
            "Values": "metric_value_keur",
        }
    )
    raw["customer_code"] = raw["customer_code"].map(lambda value: normalize_code(value, 8))
    raw["brand_code"] = raw["brand_code"].map(lambda value: normalize_code(value).upper())
    raw["pack_type"] = raw["pack_type"].map(lambda value: normalize_code(value).upper())
    raw["pack_size"] = raw["pack_size"].map(lambda value: normalize_code(value).upper())
    raw["product_key"] = raw["pack_type"] + "|" + raw["pack_size"]
    raw["metric_code"] = raw["metric_label"].map(KPI_MAP)
    raw["metric_value_keur"] = pd.to_numeric(raw["metric_value_keur"], errors="coerce")
    raw["reporting_year"] = pd.to_numeric(raw["reporting_year"], errors="coerce").astype("Int64")
    raw["month_number"] = raw["month_label"].astype(str).str[:3].str.upper().map(MONTH_NUMBERS)
    raw["reporting_month"] = pd.to_datetime(
        {"year": raw["reporting_year"], "month": raw["month_number"], "day": 1},
        errors="coerce",
    )
    return raw, files


def load_mappings(project_root: Path) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    mapping_path = project_root / "inputs" / "mapping" / "Master Mapping.xlsx"
    workbook = load_workbook(mapping_path, data_only=False, read_only=True)
    customer_sheet = workbook["Customer Mapping"]
    brand_sheet = workbook["Brand Mapping"]
    product_sheet = workbook["Product Mapping"]

    customers = []
    for row in range(2, customer_sheet.max_row + 1):
        raw_code = customer_sheet.cell(row, 2).value
        if isinstance(raw_code, str) and raw_code.upper().startswith("=ROW("):
            raw_code = row - 1
        customers.append(
            {
                "channel_code": normalize_code(customer_sheet.cell(row, 1).value).upper(),
                "customer_code": normalize_code(raw_code, 8),
                "customer_name": normalize_code(customer_sheet.cell(row, 3).value),
            }
        )

    brands = [
        {
            "brand_code": normalize_code(brand_sheet.cell(row, 1).value).upper(),
            "brand_name": normalize_code(brand_sheet.cell(row, 2).value).upper(),
        }
        for row in range(2, brand_sheet.max_row + 1)
    ]

    product_source: list[tuple[str, str, str]] = []
    lv2_counts: dict[tuple[str, str], int] = {}
    for row in range(2, product_sheet.max_row + 1):
        lv1 = normalize_code(product_sheet.cell(row, 1).value).upper()
        lv2 = normalize_code(product_sheet.cell(row, 2).value).upper()
        lv3 = normalize_code(product_sheet.cell(row, 3).value)
        product_source.append((lv1, lv2, lv3))
        lv2_counts[(lv1, lv2)] = lv2_counts.get((lv1, lv2), 0) + 1

    products = []
    for lv1, lv2, lv3 in product_source:
        parts = lv3.split(maxsplit=1)
        pack_type = parts[0].upper() if parts else ""
        pack_size = parts[1].upper() if len(parts) > 1 else ""
        products.append(
            {
                "product_key": f"{pack_type}|{pack_size}",
                "product_name": lv3.upper(),
                "product_group_lv1": lv1,
                "product_group_lv2": lv2,
                "mapping_review_required": lv1 == "BAR"
                or (lv1 == "BOX" and lv2_counts[(lv1, lv2)] > 1),
            }
        )

    workbook.close()
    return pd.DataFrame(customers), pd.DataFrame(brands), pd.DataFrame(products)


def _join_unique(values: pd.Series) -> str:
    return ";".join(sorted({str(value) for value in values.dropna() if str(value).strip()}))


def _join_rows(values: pd.Series) -> str:
    return ";".join(str(value) for value in sorted({int(value) for value in values.dropna()}))


def pivot_sales_metrics(raw_sales: pd.DataFrame) -> pd.DataFrame:
    unsupported = raw_sales.loc[~raw_sales["metric_code"].isin(["GSV", "TURNOVER"]), "metric_label"]
    if not unsupported.empty:
        raise ValueError(f"Unsupported Sales KPI labels: {sorted(unsupported.dropna().unique())}")

    grouped = (
        raw_sales.groupby([*GRAIN_COLUMNS, "metric_code"], dropna=False)
        .agg(
            metric_value_keur=("metric_value_keur", "sum"),
            record_count=("metric_value_keur", "size"),
            source_files=("source_file", _join_unique),
            source_rows=("source_row", _join_rows),
        )
        .reset_index()
    )
    values = grouped.pivot(index=GRAIN_COLUMNS, columns="metric_code", values="metric_value_keur")
    counts = grouped.pivot(index=GRAIN_COLUMNS, columns="metric_code", values="record_count")
    files = grouped.pivot(index=GRAIN_COLUMNS, columns="metric_code", values="source_files")
    rows = grouped.pivot(index=GRAIN_COLUMNS, columns="metric_code", values="source_rows")

    values = values.rename(columns={"GSV": "gsv_keur", "TURNOVER": "turnover_keur"})
    counts = counts.rename(columns={"GSV": "gsv_record_count", "TURNOVER": "turnover_record_count"})
    files = files.rename(columns={"GSV": "gsv_source_files", "TURNOVER": "turnover_source_files"})
    rows = rows.rename(columns={"GSV": "gsv_source_rows", "TURNOVER": "turnover_source_rows"})
    wide = values.join(counts, how="outer").join(files, how="outer").join(rows, how="outer").reset_index()

    for column in ["gsv_record_count", "turnover_record_count"]:
        if column not in wide:
            wide[column] = 0
        wide[column] = wide[column].fillna(0).astype(int)
    for column in ["gsv_source_files", "turnover_source_files", "gsv_source_rows", "turnover_source_rows"]:
        if column not in wide:
            wide[column] = ""
        wide[column] = wide[column].fillna("")
    return wide


def _quality_flags(row: pd.Series) -> str:
    flags: list[str] = []
    for metric in ["gsv", "turnover"]:
        count = row[f"{metric}_record_count"]
        if count == 0:
            flags.append(f"{metric.upper()}_MISSING")
        elif count > 1:
            flags.append(f"{metric.upper()}_DUPLICATE")
    if pd.notna(row["gsv_keur"]) and float(row["gsv_keur"]) < 0:
        flags.append("GSV_NEGATIVE")
    if pd.notna(row["turnover_keur"]) and float(row["turnover_keur"]) < 0:
        flags.append("TURNOVER_NEGATIVE")
    if pd.notna(row["gsv_keur"]) and pd.notna(row["turnover_keur"]) and row["turnover_keur"] > row["gsv_keur"]:
        flags.append("TURNOVER_GT_GSV")
    for column, flag in [
        ("customer_name", "CUSTOMER_MAPPING_MISSING"),
        ("brand_name", "BRAND_MAPPING_MISSING"),
        ("product_name", "PRODUCT_MAPPING_MISSING"),
    ]:
        if pd.isna(row[column]):
            flags.append(flag)
    if bool(row.get("mapping_review_required", False)):
        flags.append("PRODUCT_MAPPING_REVIEW")
    return ";".join(flags) if flags else "OK"


def build_sales_master(project_root: Path, output_path: Path | None = None) -> dict[str, Any]:
    root = project_root.resolve()
    raw_sales, source_files = load_customer_sales(root)
    customers, brands, products = load_mappings(root)
    master = pivot_sales_metrics(raw_sales)
    master = master.merge(customers, on="customer_code", how="left", validate="many_to_one")
    master = master.merge(brands, on="brand_code", how="left", validate="many_to_one")
    master = master.merge(products, on="product_key", how="left", validate="many_to_one")

    master["reporting_year"] = master["reporting_month"].dt.year.astype("Int64")
    master["month_number"] = master["reporting_month"].dt.month.astype("Int64")
    master["discount_keur"] = master["gsv_keur"] - master["turnover_keur"]
    master["discount_pct_to"] = master["discount_keur"] / master["turnover_keur"].replace(0, np.nan)
    master["discount_pct_gsv"] = master["discount_keur"] / master["gsv_keur"].replace(0, np.nan)
    master["mapping_review_required"] = master["mapping_review_required"].fillna(False).astype(bool)
    master["data_quality_flags"] = master.apply(_quality_flags, axis=1)

    blocking = (
        r"GSV_MISSING|TURNOVER_MISSING|GSV_DUPLICATE|TURNOVER_DUPLICATE|TURNOVER_GT_GSV|"
        r"CUSTOMER_MAPPING_MISSING|BRAND_MAPPING_MISSING|PRODUCT_MAPPING_MISSING"
    )
    master["certified_for_analysis"] = ~master["data_quality_flags"].str.contains(blocking, regex=True)
    master["data_quality_status"] = np.select(
        [
            master["certified_for_analysis"] & master["data_quality_flags"].eq("OK"),
            master["certified_for_analysis"],
        ],
        ["VALID", "REVIEW"],
        default="INVALID",
    )

    columns = [
        "reporting_month", "reporting_year", "month_number",
        "customer_code", "customer_name", "channel_code",
        "brand_code", "brand_name",
        "product_key", "product_name", "pack_type", "pack_size",
        "product_group_lv1", "product_group_lv2",
        "gsv_keur", "turnover_keur", "discount_keur", "discount_pct_to", "discount_pct_gsv",
        "gsv_record_count", "turnover_record_count", "mapping_review_required",
        "certified_for_analysis", "data_quality_status", "data_quality_flags",
        "gsv_source_files", "gsv_source_rows", "turnover_source_files", "turnover_source_rows",
    ]
    master = master[columns].sort_values(
        ["reporting_month", "customer_code", "brand_code", "product_key"], kind="stable"
    )
    master["reporting_month"] = master["reporting_month"].dt.strftime("%Y-%m-%d")

    target = (output_path or root / "outputs" / "sales_master.csv").resolve()
    if target.is_relative_to(root / "inputs"):
        raise ValueError("Sales master output cannot overwrite any input file.")
    target.parent.mkdir(parents=True, exist_ok=True)
    temporary = target.with_suffix(target.suffix + ".tmp")
    master.to_csv(temporary, index=False, encoding="utf-8")
    temporary.replace(target)

    return {
        "output_path": target.as_posix(),
        "source_file_count": len(source_files),
        "source_record_count": int(len(raw_sales)),
        "master_row_count": int(len(master)),
        "certified_row_count": int(master["certified_for_analysis"].sum()),
        "review_row_count": int((master["data_quality_status"] == "REVIEW").sum()),
        "invalid_row_count": int((master["data_quality_status"] == "INVALID").sum()),
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Build the UniSweet Sales master CSV.")
    parser.add_argument(
        "--project-root",
        type=Path,
        default=Path(__file__).resolve().parents[1],
        help="Project root containing inputs/sales and inputs/mapping.",
    )
    parser.add_argument("--output", type=Path, default=None, help="Optional output CSV path.")
    args = parser.parse_args()
    print(json.dumps(build_sales_master(args.project_root, args.output), indent=2, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
