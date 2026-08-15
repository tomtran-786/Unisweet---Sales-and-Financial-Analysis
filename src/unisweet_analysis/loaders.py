from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass, asdict
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from unisweet_analysis.config import ProjectPaths, parse_market_period


@dataclass(frozen=True)
class SourceFile:
    source_type: str
    path: str
    rows: int
    sha256: str
    modified_at_utc: str
    schema_status: str
    note: str


def _hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _source(
    paths: ProjectPaths,
    path: Path,
    source_type: str,
    rows: int,
    valid: bool,
    note: str,
) -> SourceFile:
    modified = datetime.fromtimestamp(path.stat().st_mtime, timezone.utc).isoformat()
    return SourceFile(
        source_type=source_type,
        path=path.relative_to(paths.root).as_posix(),
        rows=int(rows),
        sha256=_hash(path),
        modified_at_utc=modified,
        schema_status="PASS" if valid else "FAIL",
        note=note,
    )


def normalize_code(value: Any, width: int | None = None) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
    else:
        text = str(value).strip()
    return text.zfill(width) if width and text.isdigit() else text


def load_sales(
    paths: ProjectPaths,
    settings: dict[str, Any],
) -> tuple[pd.DataFrame, list[SourceFile]]:
    required = settings["sales_required_columns"]

    def sort_key(path: Path) -> tuple[int, int | str]:
        match = re.fullmatch(r"Cust\s+(\d+)", path.stem, re.IGNORECASE)
        return (0, int(match.group(1))) if match else (1, path.name)

    files = sorted(
        [path for path in paths.sales_dir.glob("Cust *.xlsx") if not path.name.startswith("~$")],
        key=sort_key,
    )
    frames: list[pd.DataFrame] = []
    sources: list[SourceFile] = []
    for path in files:
        frame = pd.read_excel(path, engine="openpyxl", dtype=object)
        valid_name = re.fullmatch(r"Cust\s+(\d+)", path.stem, re.IGNORECASE) is not None
        valid_schema = frame.columns.tolist() == required and valid_name
        sources.append(
            _source(
                paths,
                path,
                "sales",
                len(frame),
                valid_schema,
                "OK" if valid_schema else f"Expected columns {required}; received {frame.columns.tolist()}",
            )
        )
        if not valid_schema:
            continue
        frame = frame.dropna(how="all").copy()
        frame["source_file"] = path.relative_to(paths.root).as_posix()
        frame["source_row"] = frame.index + 2
        frames.append(frame)
    if not frames:
        return pd.DataFrame(), sources
    raw = pd.concat(frames, ignore_index=True)
    raw = raw.rename(
        columns={
            "Customer Code": "customer_code",
            "Brand Code": "brand_code",
            "Pack Type": "pack_type",
            "Pack Size": "pack_size",
            "Month": "month_label",
            "Year": "reporting_year",
            "KPI": "metric_label",
            "Values": "metric_value_keur",
        }
    )
    raw["customer_code"] = raw["customer_code"].map(lambda value: normalize_code(value, 8))
    raw["brand_code"] = raw["brand_code"].map(lambda value: normalize_code(value).upper())
    raw["pack_type"] = raw["pack_type"].map(lambda value: normalize_code(value).upper())
    raw["pack_size"] = raw["pack_size"].map(lambda value: normalize_code(value).upper())
    raw["product_key"] = raw["pack_type"] + "|" + raw["pack_size"]
    raw["metric_code"] = raw["metric_label"].map(settings["sales_kpis"])
    raw["metric_value_keur"] = pd.to_numeric(raw["metric_value_keur"], errors="coerce")
    raw["reporting_year"] = pd.to_numeric(raw["reporting_year"], errors="coerce").astype("Int64")
    month_numbers = {
        month: number
        for number, month in enumerate(
            ["JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"],
            start=1,
        )
    }
    raw["month_number"] = raw["month_label"].astype(str).str[:3].str.upper().map(
        month_numbers
    )
    raw["reporting_month"] = pd.to_datetime(
        {"year": raw["reporting_year"], "month": raw["month_number"], "day": 1},
        errors="coerce",
    )
    return raw, sources


def load_mappings(
    paths: ProjectPaths,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, SourceFile]:
    workbook = load_workbook(paths.mapping_file, data_only=False, read_only=True)
    customer_sheet = workbook["Customer Mapping"]
    brand_sheet = workbook["Brand Mapping"]
    product_sheet = workbook["Product Mapping"]

    customers = []
    for row in range(2, customer_sheet.max_row + 1):
        raw_code = customer_sheet.cell(row, 2).value
        if isinstance(raw_code, str) and raw_code.upper().startswith("=ROW("):
            raw_code = row - 1
        customers.append(
            {
                "channel_code": normalize_code(customer_sheet.cell(row, 1).value).upper(),
                "customer_code": normalize_code(raw_code, 8),
                "customer_name": normalize_code(customer_sheet.cell(row, 3).value),
            }
        )

    brands = [
        {
            "brand_code": normalize_code(brand_sheet.cell(row, 1).value).upper(),
            "brand_name": normalize_code(brand_sheet.cell(row, 2).value).upper(),
        }
        for row in range(2, brand_sheet.max_row + 1)
    ]

    products_source: list[tuple[str, str, str]] = []
    lv2_counts: dict[tuple[str, str], int] = {}
    for row in range(2, product_sheet.max_row + 1):
        lv1 = normalize_code(product_sheet.cell(row, 1).value).upper()
        lv2 = normalize_code(product_sheet.cell(row, 2).value).upper()
        lv3 = normalize_code(product_sheet.cell(row, 3).value)
        products_source.append((lv1, lv2, lv3))
        lv2_counts[(lv1, lv2)] = lv2_counts.get((lv1, lv2), 0) + 1
    products = []
    for lv1, lv2, lv3 in products_source:
        parts = lv3.split(maxsplit=1)
        pack_type = parts[0].upper() if parts else ""
        pack_size = parts[1].upper() if len(parts) > 1 else ""
        review = lv1 == "BAR" or (lv1 == "BOX" and lv2_counts[(lv1, lv2)] > 1)
        products.append(
            {
                "product_key": f"{pack_type}|{pack_size}",
                "pack_type": pack_type,
                "pack_size": pack_size,
                "product_group_lv1": lv1,
                "product_group_lv2": lv2,
                "product_name": lv3.upper(),
                "mapping_review_required": review,
            }
        )
    row_count = len(customers) + len(brands) + len(products)
    workbook.close()
    source = _source(paths, paths.mapping_file, "mapping", row_count, True, "OK")
    return pd.DataFrame(customers), pd.DataFrame(brands), pd.DataFrame(products), source


def load_pnl(paths: ProjectPaths) -> tuple[pd.DataFrame, SourceFile]:
    values_book = load_workbook(paths.pnl_file, data_only=True, read_only=True)
    formulas_book = load_workbook(paths.pnl_file, data_only=False, read_only=True)
    values = values_book["PnL table"]
    formulas = formulas_book["PnL table"]
    metric_aliases = {
        "GROSS SALES VALUE (GSV)": "gsv_keur",
        "DISCOUNT": "discount_keur",
        "TURNOVER (TO)": "turnover_keur",
        "TOTAL SUPPLY CHAIN COST": "supply_chain_cost_keur",
        "GROSS PROFIT (GP)": "gross_profit_keur",
        "MARKETING EXPENSE": "marketing_expense_keur",
        "PROFIT BEFORE OVERHEADS (PBO)": "pbo_keur",
    }
    metric_rows: dict[int, str] = {}
    for row in range(1, min(values.max_row, 20) + 1):
        label = normalize_code(values.cell(row, 1).value).upper()
        if label in metric_aliases:
            metric_rows[row] = metric_aliases[label]

    columns: list[tuple[str, int, int]] = []
    active_brand = ""
    for column in range(2, values.max_column + 1):
        brand = normalize_code(values.cell(1, column).value).upper()
        if brand:
            active_brand = brand
        year = values.cell(2, column).value
        if isinstance(year, (int, float)) and float(year).is_integer():
            columns.append((active_brand, int(year), column))

    rows = []
    for brand, year, column in columns:
        if brand == "TOTAL":
            continue
        record: dict[str, Any] = {"brand_name": brand, "reporting_year": year}
        for row, metric in metric_rows.items():
            record[metric] = values.cell(row, column).value
            record[f"{metric}_formula"] = formulas.cell(row, column).value
            record[f"{metric}_cell"] = f"{get_column_letter(column)}{row}"
        rows.append(record)
    expected = set(metric_aliases.values())
    valid = bool(columns) and set(metric_rows.values()) == expected
    values_book.close()
    formulas_book.close()
    source = _source(
        paths,
        paths.pnl_file,
        "pnl",
        len(rows),
        valid,
        "OK" if valid else "Missing governed P&L metrics or Brand × Year columns",
    )
    return pd.DataFrame(rows), source


def load_market(paths: ProjectPaths) -> tuple[pd.DataFrame, SourceFile]:
    parsed = parse_market_period(paths.market_file)
    workbook = load_workbook(paths.market_file, data_only=True, read_only=True)
    sheet = workbook["Market Data"]
    rows = []
    if parsed:
        year, month = parsed
        reporting_date = date(year, month, 1)
        reporting_period = f"MAT {reporting_date.strftime('%b')} {year}"
    else:
        reporting_date = None
        reporting_period = ""
    for row in range(5, sheet.max_row + 1):
        channel, segment, manufacturer, brand = [sheet.cell(row, column).value for column in range(2, 6)]
        if channel is None and manufacturer is None:
            continue
        values = [sheet.cell(row, column).value for column in range(6, 12)]
        manufacturer_code = normalize_code(manufacturer).upper()
        brand_code = normalize_code(brand).upper()
        if manufacturer_code == "CATEGORY":
            row_type = "CATEGORY_TOTAL"
        elif not brand_code:
            row_type = "MANUFACTURER_TOTAL"
        else:
            row_type = "BRAND"
        rows.append(
            {
                "reporting_period": reporting_period,
                "reporting_date": reporting_date,
                "channel_code": normalize_code(channel).upper(),
                "segment_name": normalize_code(segment),
                "manufacturer_name": manufacturer_code,
                "brand_name": brand_code,
                "row_type": row_type,
                "sales_value_mat_1_meur": values[0],
                "sales_value_mat_meur": values[1],
                "sales_value_movement_meur": values[2],
                "value_share_mat_1": float(values[3]) / 100 if values[3] is not None else None,
                "value_share_mat": float(values[4]) / 100 if values[4] is not None else None,
                "share_movement_pp": values[5],
                "source_row": row,
            }
        )
    workbook.close()
    valid = parsed is not None and len(rows) > 0
    source = _source(
        paths,
        paths.market_file,
        "market",
        len(rows),
        valid,
        "OK" if valid else "Market filename must contain a valid MAT period",
    )
    return pd.DataFrame(rows), source


def source_records(sources: list[SourceFile]) -> list[dict[str, Any]]:
    return [asdict(source) for source in sources]
